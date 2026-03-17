"""
Script untuk mengambil data jurnal Perguruan Tinggi dari SINTA
menggunakan requests + BeautifulSoup (tanpa browser)

Sumber: https://sinta.kemdiktisaintek.go.id/

Alur:
  1. Baca outs/namapt_list.json  -> daftar PT dengan field 'kode' (kodept)
  2. Cari PT di SINTA via  /affiliations?q=[kodept]
     Ambil SINTA ID dari URL /affiliations/profile/{sinta_id}
  3. Ambil profil PT dari /affiliations/profile/{sinta_id}
  4. Ambil semua jurnal dari /journals/index/{sinta_id}?page=N (loop halaman)
  5. Simpan ke outs/sinta_jurnal.json dengan index kodept

Output: outs/sinta_jurnal.json
  {
    "011003": {
      "sinta_id": "581",
      "profil_pt": {
        "nama_pt": "...", "singkatan": "...", "lokasi": "...",
        "id_sinta": "...", "kode_pt": "...",
        "authors": "626", "departments": "47", "journals": "36",
        "sinta_score_overall": "648.451", ...
      },
      "jurnal": [
        {
          "nama_jurnal": "...", "sinta_url": "...",
          "url_google_scholar": "...", "url_website": "...", "url_editor": "...",
          "p_issn": "...", "e_issn": "...", "subject_area": "...",
          "akreditasi": "S2", "garuda_indexed": true, "url_garuda": "...",
          "impact": "42,14", "h5_index": "49",
          "citations_5yr": "6.210", "citations": "6.266"
        }, ...
      ]
    }, ...
  }
"""

import argparse
import json
import os
import re
import time

import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

BASE_DIR    = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
INPUT_FILE  = os.path.join(BASE_DIR, "utils/outs/namapt_list.json")
OUTPUT_FILE  = os.path.join(BASE_DIR, "utils/outs/sinta_jurnal.json")
OUTPUT_EXCEL = os.path.join(BASE_DIR, "utils/outs/ptma_jurnal_sinta.xlsx")

SINTA_BASE    = "https://sinta.kemdiktisaintek.go.id"
SEARCH_URL    = SINTA_BASE + "/affiliations?q={kode}"
PROFILE_URL   = SINTA_BASE + "/affiliations/profile/{sinta_id}"
JOURNALS_URL  = SINTA_BASE + "/journals/index/{sinta_id}?page={page}"

BETWEEN_REQUEST = 1.5  # detik jeda antar request
BETWEEN_PT      = 2    # detik jeda antar PT
REQUEST_TIMEOUT = 30   # detik timeout per request
RETRY_COUNT     = 4    # jumlah retry jika gagal
RETRY_WAIT      = 8    # detik jeda antar retry

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (X11; Linux x86_64; rv:124.0) Gecko/20100101 Firefox/124.0"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "id,en-US;q=0.7,en;q=0.3",
    "Referer": SINTA_BASE,
}


# ---------------------------------------------------------------------------
# HTTP helper
# ---------------------------------------------------------------------------

def make_session():
    s = requests.Session()
    s.headers.update(HEADERS)
    return s


def fetch(session, url, retries=RETRY_COUNT):
    """GET url, kembalikan BeautifulSoup atau None jika semua retry gagal."""
    for attempt in range(1, retries + 1):
        try:
            r = session.get(url, timeout=REQUEST_TIMEOUT)
            r.raise_for_status()
            return BeautifulSoup(r.text, "lxml")
        except Exception as e:
            print(f"  [attempt {attempt}/{retries}] Error {url}: {e}")
            if attempt < retries:
                time.sleep(RETRY_WAIT)
    return None


# ---------------------------------------------------------------------------
# Cari SINTA ID dari pencarian kodept
# ---------------------------------------------------------------------------

def find_sinta_id(session, kode):
    """
    Cari PT di SINTA dengan query kodept.
    Return SINTA ID (string angka) atau None.
    """
    url = SEARCH_URL.format(kode=kode)
    soup = fetch(session, url)
    if soup is None:
        return None

    # Cari href /affiliations/profile/{id}
    for a in soup.find_all("a", href=True):
        m = re.search(r"/affiliations/profile/(\d+)", a["href"])
        if m:
            sinta_id = m.group(1)
            print(f"  SINTA ID: {sinta_id}")
            return sinta_id

    print(f"  SINTA ID tidak ditemukan untuk kode: {kode}")
    return None


# ---------------------------------------------------------------------------
# Profil PT
# ---------------------------------------------------------------------------

def scrape_profil_pt(session, sinta_id):
    """
    Ambil data profil PT dari /affiliations/profile/{sinta_id}.
    Field:
      .affil-abbrev           -> singkatan
      .affil-loc              -> lokasi
      .affil-code             -> id_sinta, kode_pt
      div.stat-num/.stat-text -> authors, departments, journals
      div.pr-num/.pr-txt      -> sinta_score_*
    """
    url = PROFILE_URL.format(sinta_id=sinta_id)
    soup = fetch(session, url)
    if soup is None:
        return {"sinta_profile_url": url}

    profil = {"sinta_profile_url": url}

    # Nama PT
    for sel in [("div", "univ-name"), None]:
        try:
            univ_div = soup.find("div", class_="univ-name")
            h = univ_div.find(["h3", "h2"]) if univ_div else soup.find("h3")
            if h:
                profil["nama_pt"] = h.get_text(strip=True)
                break
        except Exception:
            pass

    # Singkatan  (.affil-abbrev)
    el = soup.find(class_="affil-abbrev")
    if el:
        profil["singkatan"] = el.get_text(strip=True)

    # Lokasi  (.affil-loc)
    el = soup.find(class_="affil-loc")
    if el:
        profil["lokasi"] = el.get_text(strip=True)

    # ID SINTA dan kode PT  (.affil-code)
    # Contoh teks: "ID : 581  CODE : 011003"
    el = soup.find(class_="affil-code")
    if el:
        code_text = el.get_text(" ", strip=True)
        m = re.search(r"ID\s*:\s*(\d+)", code_text)
        if m:
            profil["id_sinta"] = m.group(1)
        m = re.search(r"CODE\s*:\s*(\S+)", code_text)
        if m:
            profil["kode_pt"] = m.group(1)

    # Stat cards: Authors, Departments, Journals
    stat_nums  = soup.find_all("div", class_="stat-num")
    stat_texts = soup.find_all("div", class_="stat-text")
    for num_el, txt_el in zip(stat_nums, stat_texts):
        num = num_el.get_text(strip=True)
        lbl = txt_el.get_text(strip=True)
        if num and lbl:
            profil[lbl.lower().replace(" ", "_")] = num

    # SINTA Score: pr-num / pr-txt
    pr_nums = soup.find_all("div", class_="pr-num")
    pr_txts = soup.find_all("div", class_="pr-txt")
    for num_el, txt_el in zip(pr_nums, pr_txts):
        num = num_el.get_text(strip=True)
        lbl = txt_el.get_text(strip=True)
        if num and lbl:
            profil[lbl.lower().replace(" ", "_")] = num

    return profil


# ---------------------------------------------------------------------------
# Parse satu item jurnal
# ---------------------------------------------------------------------------

def parse_journal_item(item):
    """
    Parse satu tag div.list-item menjadi dict data jurnal.
    """
    j = {}

    # Nama jurnal + URL profil SINTA
    affil_name = item.find(class_="affil-name")
    if affil_name:
        a = affil_name.find("a")
        if a:
            j["nama_jurnal"] = a.get_text(strip=True)
            j["sinta_url"]   = a.get("href", "")

    # Link eksternal: Google Scholar, Website, Editor URL
    affil_abbrev = item.find(class_="affil-abbrev")
    if affil_abbrev:
        for a in affil_abbrev.find_all("a", href=True):
            text = a.get_text(strip=True)
            href = a["href"]
            if "Google Scholar" in text:
                j["url_google_scholar"] = href
            elif "Editor" in text:
                j["url_editor"] = href
            elif "Website" in text:
                j["url_website"] = href

    # P-ISSN, E-ISSN, Subject Area
    # Teks: "P-ISSN : 16937619 | E-ISSN : 25804170  Subject Area : Science"
    profile_id = item.find(class_="profile-id")
    if profile_id:
        pid_text = profile_id.get_text(" ", strip=True)

        m = re.search(r"P-ISSN\s*:\s*([\w]+)", pid_text)
        j["p_issn"] = m.group(1) if m else ""

        m = re.search(r"E-ISSN\s*:\s*([\w]+)", pid_text)
        j["e_issn"] = m.group(1) if m else ""

        m = re.search(r"Subject Area\s*:\s*(.+?)(?:\s{2,}|$)", pid_text)
        j["subject_area"] = m.group(1).strip() if m else ""

    # Akreditasi + Garuda Indexed
    stat_prev = item.find(class_="stat-prev")
    if stat_prev:
        # Akreditasi: elemen dengan class "accredited"
        akred_el = stat_prev.find(class_="accredited")
        if akred_el:
            akred_text = akred_el.get_text(strip=True)
            m = re.search(r"(S\d+)", akred_text)
            j["akreditasi"] = m.group(1) if m else akred_text
        else:
            j["akreditasi"] = ""

        # Garuda Indexed
        garuda_el = stat_prev.find(class_="garuda-indexed")
        if garuda_el:
            j["garuda_indexed"] = True
            # URL Garuda dari <a> pembungkus
            parent_a = garuda_el.find_parent("a")
            j["url_garuda"] = parent_a["href"] if parent_a and parent_a.get("href") else ""
        else:
            j["garuda_indexed"] = False
            j["url_garuda"] = ""

    # Statistik: Impact, H5-index, Citations 5yr, Citations
    stat_prof = item.find(class_="stat-profile")
    if stat_prof:
        pr_nums = stat_prof.find_all("div", class_="pr-num")
        pr_txts = stat_prof.find_all("div", class_="pr-txt")
        for num_el, txt_el in zip(pr_nums, pr_txts):
            num = num_el.get_text(strip=True)
            lbl = txt_el.get_text(strip=True)
            if num and lbl:
                key = lbl.lower().replace("-", "_").replace(" ", "_")
                j[key] = num

    return j


# ---------------------------------------------------------------------------
# Ambil semua jurnal PT (semua halaman)
# ---------------------------------------------------------------------------

def scrape_journals_for_pt(session, sinta_id):
    """
    Loop halaman /journals/index/{sinta_id}?page=N sampai tidak ada item.
    Return list of dict.
    """
    all_journals = []
    page = 1

    while True:
        url = JOURNALS_URL.format(sinta_id=sinta_id, page=page)
        print(f"    Halaman {page}: {url}")

        soup = fetch(session, url)
        if soup is None:
            print(f"    Gagal fetch halaman {page}, stop")
            break

        items = soup.find_all("div", class_="list-item")
        print(f"    Ditemukan {len(items)} item")

        if not items:
            break

        for item in items:
            j = parse_journal_item(item)
            if j.get("nama_jurnal"):
                all_journals.append(j)

        # Cek apakah ada halaman berikutnya
        next_page = page + 1
        has_next = any(
            f"page={next_page}" in (a.get("href", ""))
            for a in soup.select(".pagination a")
        )
        if not has_next:
            break

        page = next_page
        time.sleep(BETWEEN_REQUEST)

    return all_journals


# ---------------------------------------------------------------------------
# Scrape satu PT
# ---------------------------------------------------------------------------

def scrape_pt(session, kode, nama):
    print(f"\n[{kode}] {nama}")

    sinta_id = find_sinta_id(session, kode)
    if not sinta_id:
        return {"sinta_id": None, "error": "PT tidak ditemukan di SINTA", "jurnal": []}

    time.sleep(BETWEEN_REQUEST)
    profil = scrape_profil_pt(session, sinta_id)

    time.sleep(BETWEEN_REQUEST)
    jurnal_list = scrape_journals_for_pt(session, sinta_id)

    print(f"  Total jurnal: {len(jurnal_list)}")

    return {
        "sinta_id":  sinta_id,
        "profil_pt": profil,
        "jurnal":    jurnal_list,
    }


# ---------------------------------------------------------------------------
# Export Excel
# ---------------------------------------------------------------------------

# Urutan dan label kolom sheet Jurnal
JURNAL_COLS = [
    ("kodept",              "Kode PT"),
    ("nama_pt",             "Nama PT"),
    ("singkatan",           "Singkatan PT"),
    ("sinta_id",            "SINTA ID PT"),
    ("nama_jurnal",         "Nama Jurnal"),
    ("sinta_url",           "URL SINTA Jurnal"),
    ("p_issn",              "P-ISSN"),
    ("e_issn",              "E-ISSN"),
    ("subject_area",        "Subject Area"),
    ("akreditasi",          "Akreditasi"),
    ("garuda_indexed",      "Garuda Indexed"),
    ("url_garuda",          "URL Garuda"),
    ("impact",              "Impact"),
    ("h5_index",            "H5-Index"),
    ("citations_5yr",       "Citations 5yr"),
    ("citations",           "Citations"),
    ("url_google_scholar",  "URL Google Scholar"),
    ("url_website",         "URL Website"),
    ("url_editor",          "URL Editor"),
]

# Urutan dan label kolom sheet Profil PT
PROFIL_COLS = [
    ("kodept",                      "Kode PT"),
    ("sinta_id",                    "SINTA ID"),
    ("nama_pt",                     "Nama PT"),
    ("singkatan",                   "Singkatan"),
    ("lokasi",                      "Lokasi"),
    ("kode_pt",                     "Kode (SINTA)"),
    ("authors",                     "Jumlah Authors"),
    ("departments",                 "Jumlah Departments"),
    ("journals",                    "Jumlah Journals"),
    ("sinta_score_overall",         "SINTA Score Overall"),
    ("sinta_score_3yr",             "SINTA Score 3Yr"),
    ("sinta_score_productivity",    "SINTA Score Productivity"),
    ("sinta_score_productivity_3yr","SINTA Score Productivity 3Yr"),
    ("sinta_profile_url",           "URL Profil SINTA"),
]


def _style_header(ws, row, ncols):
    """Beri style baris header: bold, background biru tua, teks putih."""
    header_fill = PatternFill("solid", fgColor="1F497D")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="FFFFFF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill   = header_fill
        cell.font   = header_font
        cell.alignment = header_align
        cell.border = border


def _style_data_row(ws, row, ncols, even):
    """Warna baris data bergantian."""
    fill = PatternFill("solid", fgColor="DCE6F1" if even else "FFFFFF")
    align = Alignment(vertical="center", wrap_text=False)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill      = fill
        cell.alignment = align
        cell.border    = border


def _autofit(ws, cols):
    """Set lebar kolom berdasarkan panjang konten (estimasi)."""
    for i, (_, label) in enumerate(cols, 1):
        max_len = len(label)
        for row in ws.iter_rows(min_row=2, min_col=i, max_col=i):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 60)


def export_excel(json_file=OUTPUT_FILE, xlsx_file=OUTPUT_EXCEL):
    """
    Baca sinta_jurnal.json dan ekspor ke Excel dengan dua sheet:
      - Jurnal  : satu baris per jurnal, dilengkapi kolom identitas PT
      - Profil PT: satu baris per PT
    """
    print(f"\nMemuat data: {json_file}")
    with open(json_file, encoding="utf-8") as f:
        data = json.load(f)

    wb = openpyxl.Workbook()

    # ---- Sheet 1: Jurnal ----
    ws_j = wb.active
    ws_j.title = "Jurnal"
    ws_j.freeze_panes = "A2"

    # Header
    headers_j = [label for _, label in JURNAL_COLS]
    for col, label in enumerate(headers_j, 1):
        ws_j.cell(row=1, column=col, value=label)
    _style_header(ws_j, 1, len(JURNAL_COLS))

    row_j = 2
    for kodept, pt_data in data.items():
        profil   = pt_data.get("profil_pt", {})
        sinta_id = pt_data.get("sinta_id", "")
        nama_pt  = profil.get("nama_pt", "")
        singkat  = profil.get("singkatan", "")

        for jurnal in pt_data.get("jurnal", []):
            row_data = {
                "kodept":   kodept,
                "nama_pt":  nama_pt,
                "singkatan": singkat,
                "sinta_id": sinta_id,
                **jurnal,
            }
            for col, (key, _) in enumerate(JURNAL_COLS, 1):
                val = row_data.get(key, "")
                if isinstance(val, bool):
                    val = "Ya" if val else "Tidak"
                ws_j.cell(row=row_j, column=col, value=val)
            _style_data_row(ws_j, row_j, len(JURNAL_COLS), row_j % 2 == 0)
            row_j += 1

    _autofit(ws_j, JURNAL_COLS)
    ws_j.row_dimensions[1].height = 30

    # ---- Sheet 2: Profil PT ----
    ws_p = wb.create_sheet("Profil PT")
    ws_p.freeze_panes = "A2"

    headers_p = [label for _, label in PROFIL_COLS]
    for col, label in enumerate(headers_p, 1):
        ws_p.cell(row=1, column=col, value=label)
    _style_header(ws_p, 1, len(PROFIL_COLS))

    for row_p, (kodept, pt_data) in enumerate(data.items(), 2):
        profil = pt_data.get("profil_pt", {})
        row_data = {
            "kodept":  kodept,
            "sinta_id": pt_data.get("sinta_id", ""),
            **profil,
        }
        for col, (key, _) in enumerate(PROFIL_COLS, 1):
            ws_p.cell(row=row_p, column=col, value=row_data.get(key, ""))
        _style_data_row(ws_p, row_p, len(PROFIL_COLS), row_p % 2 == 0)

    _autofit(ws_p, PROFIL_COLS)
    ws_p.row_dimensions[1].height = 30

    wb.save(xlsx_file)
    total_jurnal = row_j - 2
    total_pt     = len(data)
    print(f"Excel disimpan : {xlsx_file}")
    print(f"Sheet Jurnal   : {total_jurnal} baris")
    print(f"Sheet Profil PT: {total_pt} baris")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Scrape data jurnal PT dari SINTA",
        formatter_class=argparse.RawTextHelpFormatter,
    )
    parser.add_argument(
        "--export",
        action="store_true",
        help="Export data JSON yang sudah ada ke Excel (tanpa scraping ulang)",
    )
    parser.add_argument(
        "--input",
        default=OUTPUT_FILE,
        metavar="FILE",
        help=f"File JSON input untuk --export (default: {OUTPUT_FILE})",
    )
    parser.add_argument(
        "--output",
        default=OUTPUT_EXCEL,
        metavar="FILE",
        help=f"File Excel output untuk --export (default: {OUTPUT_EXCEL})",
    )
    args = parser.parse_args()

    # Mode export saja
    if args.export:
        export_excel(json_file=args.input, xlsx_file=args.output)
        return

    # Mode scraping
    with open(INPUT_FILE, encoding="utf-8") as f:
        pt_list = json.load(f)
    print(f"Jumlah PT: {len(pt_list)}")

    # Resume: load output yang sudah ada
    if os.path.exists(OUTPUT_FILE):
        with open(OUTPUT_FILE, encoding="utf-8") as f:
            results = json.load(f)
        print(f"Resume: {len(results)} PT sudah diproses")
    else:
        results = {}

    session = make_session()

    for i, pt in enumerate(pt_list, 1):
        kode = pt.get("kode", "")
        nama = pt.get("target", pt.get("keyword", ""))

        if not kode:
            continue

        if kode in results:
            print(f"[{i}/{len(pt_list)}] Skip: {kode} {nama}")
            continue

        try:
            data = scrape_pt(session, kode, nama)
            results[kode] = data
        except Exception as e:
            print(f"  ERROR: {e}")
            results[kode] = {"sinta_id": None, "error": str(e), "jurnal": []}

        # Auto-save setelah tiap PT
        with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
            json.dump(results, f, ensure_ascii=False, indent=2)

        time.sleep(BETWEEN_PT)

    total_jurnal = sum(len(v.get("jurnal", [])) for v in results.values())
    print(f"\n=== Selesai Scraping ===")
    print(f"Total PT      : {len(results)}")
    print(f"Total jurnal  : {total_jurnal}")
    print(f"Output JSON   : {OUTPUT_FILE}")

    # Auto-export Excel setelah scraping selesai
    export_excel()


if __name__ == "__main__":
    main()
